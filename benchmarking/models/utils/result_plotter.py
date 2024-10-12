import argparse
import os

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd


def collect_data(data_file_path):
    """
    Collects the data from an excel (xlsx) file.
    The data related to inactive sheets, models, and model types will be excluded.

    Args:
        data_file_path (str): The path to the excel file.

    Returns:
        tuple: A tuple containing the data and the models metadata.
            - data (dict): The data collected from the excel file. It will be in the following format:
                {plot_name: {"color", "marker", "data_points": [{"model_name", "flash", "ram", "exe": {"n_tests", "average", "std"}, "error": {"n_tests", "average", "std"}}]}}
            - models_metadata (dict): The metadata of the models. It will be in the following format:
                {model_name: {"display_name", "parameters", "MACs"}}
    """
    def _resolve_merged_cells(file_path):
        wb = openpyxl.load_workbook(file_path)
        xls = pd.read_excel(file_path, sheet_name=None)

        resolved_dfs = {}
        for sheet_name, df in xls.items():
            sheet = wb[sheet_name]

            for merged_cell_range in sheet.merged_cells.ranges:
                row_min, col_min, row_max, col_max = merged_cell_range.min_row, merged_cell_range.min_col, merged_cell_range.max_row, merged_cell_range.max_col
                merged_value = sheet.cell(row=row_min, column=col_min).value

                if merged_value is None or row_min <= 2:    # skip empty rows and header row (first two rows)
                    continue

                for row in range(row_min, row_max + 1):
                    for col in range(col_min, col_max + 1):
                        # convert to zero-based indexing for pandas DataFrame
                        # also, first row is header row and not counted in pandas DataFrame
                        r = row - 2
                        c = col - 1
                        if pd.isna(df.iat[r, c]):
                            df.iat[r, c] = merged_value

            resolved_dfs[sheet_name] = df

        return resolved_dfs

    def _replace_nan_with_none(value):
        return None if pd.isna(value) else value

    sheets = _resolve_merged_cells(data_file_path)

    # collect the sheets metadata
    sheets_metadata = {}    # {sheet_name: {active, display_name, color}}
    for sheet_name, sheet in sheets.items():
        if sheet_name.lower() == "metadata - sheets":
            if "Sheet Name" not in sheet.columns or "Active" not in sheet.columns or "Display Name" not in sheet.columns or "Color" not in sheet.columns:
                raise ValueError("The metadata sheet should contain the columns 'Sheet Name', 'Active', 'Display Name', and 'Color'.")

            for _, row in sheet.iterrows():
                if _replace_nan_with_none(row["Sheet Name"]) is not None:
                    sheets_metadata[row["Sheet Name"]] = {
                        "active": _replace_nan_with_none(row["Active"]),
                        "display_name": _replace_nan_with_none(row["Display Name"]),
                        "color": _replace_nan_with_none(row["Color"])
                    }

    # collect the models metadata
    models_metadata = {}    # {model_name: {display_name, parameters, MACs}}
    for sheet_name, sheet in sheets.items():
        if sheet_name.lower() == "metadata - models":
            if "Model Name" not in sheet.columns or "Active" not in sheet.columns or "Display Name" not in sheet.columns or "Parameters" not in sheet.columns or "MACs" not in sheet.columns:
                raise ValueError("The metadata sheet should contain the columns 'Model Name', 'Active', 'Display Name', 'Parameters', and 'MACs'.")

            for _, row in sheet.iterrows():
                if _replace_nan_with_none(row["Model Name"]) is not None:
                    models_metadata[row["Model Name"]] = {
                        "active": _replace_nan_with_none(row["Active"]),
                        "display_name": _replace_nan_with_none(row["Display Name"]),
                        "parameters": _replace_nan_with_none(row["Parameters"]),
                        "MACs": _replace_nan_with_none(row["MACs"])
                    }

    # collect the model types metadata
    model_types_metadata = {}   # {model_type: {active, display_name, marker}}
    for sheet_name, sheet in sheets.items():
        if sheet_name.lower() == "metadata - model types":
            if "Model Type" not in sheet.columns or "Active" not in sheet.columns or "Display Name" not in sheet.columns or "Marker" not in sheet.columns:
                raise ValueError("The metadata sheet should contain the columns 'Model Type', 'Active', 'Display Name', and 'Marker'.")

            for _, row in sheet.iterrows():
                if _replace_nan_with_none(row["Model Type"]) is not None:
                    model_types_metadata[row["Model Type"]] = {
                        "active": _replace_nan_with_none(row["Active"]),
                        "display_name": _replace_nan_with_none(row["Display Name"]),
                        "marker": _replace_nan_with_none(row["Marker"])
                    }

    # collect the data
    plots = {}   # {plot_name: {model_name, flash, ram, exe: {n_tests, average, std}, error: {n_tests, average, std}}}
    for sheet_name, sheet in sheets.items():
        if sheet_name.lower() == "metadata - sheets" or sheet_name.lower() == "metadata - models" or sheet_name.lower() == "metadata - model types":
            continue
        if sheet_name in sheets_metadata and sheets_metadata[sheet_name]["active"] is not None and sheets_metadata[sheet_name]["active"] == 0:
            continue

        # check if the sheet contains the required columns
        if ("Model Name" not in sheet.columns or
            "Flash (kB)" not in sheet.columns or
            "RAM (kB)" not in sheet.columns or
            "Execution Time (ms)" not in sheet.columns or
            "Error (MAE)" not in sheet.columns):
            raise ValueError("The data sheets should contain the columns 'Model Name', 'Flash (kB)', 'RAM (kB)', 'Execution Time (ms)', and 'Error (MAE)'.")

        # find the index of the column
        columns = {}
        columns["Model Name"] = sheet.columns.get_loc("Model Name")
        columns["Flash (kB)"] = sheet.columns.get_loc("Flash (kB)")
        columns["RAM (kB)"] = sheet.columns.get_loc("RAM (kB)")
        columns["Execution Time (ms) - n tests"] = sheet.columns.get_loc("Execution Time (ms)")
        columns["Execution Time (ms) - average"] = columns["Execution Time (ms) - n tests"] + 1
        columns["Execution Time (ms) - std"] = columns["Execution Time (ms) - n tests"] + 2
        columns["Error (MAE) - n tests"] = sheet.columns.get_loc("Error (MAE)")
        columns["Error (MAE) - average"] = columns["Error (MAE) - n tests"] + 1
        columns["Error (MAE) - std"] = columns["Error (MAE) - n tests"] + 2

        # check if the second row contains the required columns
        if sheet.iloc[0, columns["Execution Time (ms) - n tests"]] != "n tests":
            raise ValueError("The Execution Time (ms) column should contain 'n tests' in the second row.")
        if sheet.iloc[0, columns["Execution Time (ms) - average"]] != "average":
            raise ValueError("The Execution Time (ms) column should contain 'average' in the second row.")
        if sheet.iloc[0, columns["Execution Time (ms) - std"]] != "std":
            raise ValueError("The Execution Time (ms) column should contain 'std' in the second row.")
        if sheet.iloc[0, columns["Error (MAE) - n tests"]] != "n tests":
            raise ValueError("The Error (MAE) column should contain 'n tests' in the second row.")
        if sheet.iloc[0, columns["Error (MAE) - average"]] != "average":
            raise ValueError("The Error (MAE) column should contain 'average' in the second row.")
        if sheet.iloc[0, columns["Error (MAE) - std"]] != "std":
            raise ValueError("The Error (MAE) column should contain 'std' in the second row.")

        # data collection loop
        for i, row in sheet.iterrows():
            if i == 0:   # skip the second row as its a part of the column names
                continue

            model_name = _replace_nan_with_none(row["Model Name"])
            if model_name is not None:
                model_type = _replace_nan_with_none(row["Model Type"])
                plot_name = sheet_name
                if sheet_name in sheets_metadata and sheets_metadata[sheet_name]["display_name"] is not None:
                    plot_name = sheets_metadata[sheet_name]["display_name"]

                if (model_name in models_metadata and
                    models_metadata[model_name]["active"] is not None and
                    models_metadata[model_name]["active"] == 0):
                    continue
                if (model_type is not None and
                    model_type in model_types_metadata and
                    model_types_metadata[model_type]["active"] is not None and
                    model_types_metadata[model_type]["active"] == 0):
                    continue

                if model_type is not None:
                    if model_type in model_types_metadata and model_types_metadata[model_type]["display_name"] is not None:
                            plot_name += " [{}]".format(model_types_metadata[model_type]["display_name"])
                    else:
                        plot_name += " [{}]".format(model_type)

                if plot_name not in plots:
                    color = None
                    if sheet_name in sheets_metadata and sheets_metadata[sheet_name]["color"] is not None:
                        color = sheets_metadata[sheet_name]["color"]
                    marker = 'o'
                    if model_type is not None and model_type in model_types_metadata and model_types_metadata[model_type]["marker"] is not None:
                        marker = model_types_metadata[model_type]["marker"]
                    plots[plot_name] = {
                        "color": color,
                        "marker": marker,
                        "data_points": []
                    }
                plots[plot_name]["data_points"].append({
                    "model_name": model_name,
                    "flash": _replace_nan_with_none(row["Flash (kB)"]),
                    "ram": _replace_nan_with_none(row["RAM (kB)"]),
                    "exe": {
                        "n_tests": _replace_nan_with_none(row.iloc[columns["Execution Time (ms) - n tests"]]),
                        "average": _replace_nan_with_none(row.iloc[columns["Execution Time (ms) - average"]]),
                        "std": _replace_nan_with_none(row.iloc[columns["Execution Time (ms) - std"]])
                    },
                    "error": {
                        "n_tests": _replace_nan_with_none(row.iloc[columns["Error (MAE) - n tests"]]),
                        "average": _replace_nan_with_none(row.iloc[columns["Error (MAE) - average"]]),
                        "std": _replace_nan_with_none(row.iloc[columns["Error (MAE) - std"]])
                    }
                })

    return plots, models_metadata


def fill_plot_gaps(ax, line, opacity=0.5):
    """
    This function fills the gap caused by the None values with a dashed line.

    Args:
        ax (matplotlib.axes.Axes): The axes where the line is plotted.
        line (matplotlib.lines.Line2D): The line to fill the gap for.
        opacity (float): The opacity of the dashed line.
    """
    # fill the gap caused by the None values
    x = line.get_xdata()
    y = line.get_ydata()
    for i in range(len(y) - 1):
        if y[i] is not None and y[i+1] is None:
            for j in range(i+1, len(y)):
                if y[j] is not None:
                    ax.plot([x[i], x[j]], [y[i], y[j]], color=line.get_color(), linestyle='dashed', alpha=opacity)
                    break


def save_light(fig, file_path):
    """
    Save the figure with a light background.
    Note: The attributes of the figure will be changed after using this function.

    Args:
        fig (matplotlib.figure.Figure): The figure to save.
        file_path (str): The path to save the figure, excluding the file extension.
    """
    fig.patch.set_facecolor('white')
    axes = fig.get_axes()
    for ax in axes:
        ax.set_facecolor('white')
        ax.xaxis.label.set_color('black')
        ax.yaxis.label.set_color('black')
        ax.title.set_color('black')
        ax.tick_params(axis='x', colors='black')
        ax.tick_params(axis='y', colors='black')
        ax.spines['bottom'].set_color('black')
        ax.spines['top'].set_color('black')
        ax.spines['right'].set_color('black')
        ax.spines['left'].set_color('black')

    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    fig.savefig(file_path + ".png", bbox_inches='tight', dpi=300)
    fig.savefig(file_path + ".pdf", bbox_inches='tight')


def save_dark(fig, file_path):
    """
    Save the figure with a dark background.
    Note: The attributes of the figure will be changed after using this function.

    Args:
        fig (matplotlib.figure.Figure): The figure to save.
        file_path (str): The path to save the figure, excluding the file extension.
    """
    fig.patch.set_facecolor((30/255, 30/255, 30/255))
    axes = fig.get_axes()
    for ax in axes:
        ax.set_facecolor((30/255, 30/255, 30/255))
        ax.xaxis.label.set_color('white')
        ax.yaxis.label.set_color('white')
        ax.title.set_color('white')
        ax.tick_params(axis='x', colors='white')
        ax.tick_params(axis='y', colors='white')
        ax.spines['bottom'].set_color('white')
        ax.spines['top'].set_color('white')
        ax.spines['right'].set_color('white')
        ax.spines['left'].set_color('white')

    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    fig.savefig(file_path + ".png", bbox_inches='tight', dpi=300)
    fig.savefig(file_path + ".pdf", bbox_inches='tight')


def main(data_file_path, save_dir):
    """
    This function plots the results of the benchmarking test.

    Args:
        data_file_path (str): The path to the excel file containing the benchmarking results.
        save_dir (str): The directory where the plots will be saved.
    """
    plots, models_metadata = collect_data(data_file_path)
    opacity = 0.8

    present_models = []
    for _, plot_vals in plots.items():
        for data_point in plot_vals["data_points"]:
            if data_point["model_name"] not in present_models:
                present_models.append(data_point["model_name"])

    display_names = []
    for model_name in present_models:
        if model_name in models_metadata and models_metadata[model_name]["display_name"] is not None:
            display_names.append(models_metadata[model_name]["display_name"])
        else:
            display_names.append(model_name)

    parameters = []
    MACs = []
    for model_name in present_models:
        if model_name in models_metadata:
            parameters.append(models_metadata[model_name]["parameters"])
            MACs.append(models_metadata[model_name]["MACs"])
        else:
            parameters.append(None)
            MACs.append(None)

    plot_params = any([param is not None for param in parameters])
    plot_MACs = any([MAC is not None for MAC in MACs])

    # plot models parameters and MACs
    if plot_params and plot_MACs:
        fig, ax1 = plt.subplots()
        ax2 = ax1.twinx()
        p1 = ax1.plot(display_names, parameters, 'tab:blue', label='Parameters', marker='o', alpha=opacity)
        p2 = ax2.plot(display_names, MACs, 'tab:orange', label='MACs', marker='o', alpha=opacity)

        fill_plot_gaps(ax1, p1[0])
        fill_plot_gaps(ax2, p2[0])

        ax1.set_yscale('log')
        ax2.set_yscale('log')

        ax1.set_xlabel('Models')
        ax1.set_ylabel('Parameters')
        ax2.set_ylabel('MACs')

        lines = [p1[0], p2[0]]
        labels = [line.get_label() for line in lines]
        ax1.legend(lines, labels, loc='best')

        ax1.grid(True, which='major', linestyle='-', linewidth=0.5)
        ax1.grid(True, which='minor', linestyle='--', linewidth=0.2)
        ax1.xaxis.set_tick_params(rotation=45)

        save_light(fig, os.path.join(save_dir, "params_MACs"))
        save_dark(fig, os.path.join(save_dir, "dark", "params_MACs"))

    elif plot_params:
        fig, ax = plt.subplots()
        line = ax.plot(display_names, parameters, 'tab:blue', label='Parameters', marker='o', alpha=opacity)
        fill_plot_gaps(ax, line[0])

        ax.set_yscale('log')
        ax.set_xlabel('Models')
        ax.set_ylabel('Parameters')

        ax.grid(True, which='major', linestyle='-', linewidth=0.5)
        ax.grid(True, which='minor', linestyle='--', linewidth=0.2)
        ax.xaxis.set_tick_params(rotation=45)

        save_light(fig, os.path.join(save_dir, "params"))
        save_dark(fig, os.path.join(save_dir, "dark", "params"))

    elif plot_MACs:
        line = ax.plot(display_names, MACs, 'tab:orange', label='MACs', marker='o', alpha=opacity)
        fill_plot_gaps(ax, line[0])

        ax.set_yscale('log')
        ax.set_xlabel('Models')
        ax.set_ylabel('MACs')

        ax.grid(True, which='major', linestyle='-', linewidth=0.5)
        ax.grid(True, which='minor', linestyle='--', linewidth=0.2)
        ax.xaxis.set_tick_params(rotation=45)

        save_light(fig, os.path.join(save_dir, "MACs"))
        save_dark(fig, os.path.join(save_dir, "dark", "MACs"))

    # collect all x points in all plots, respecting their order
    figure_x_names = []
    figure_x_display_names = []
    figure_x_parameters = []
    figure_x_MACs = []

    for plot_name, plot_vals in plots.items():
        last_found = -1
        for data_point in plot_vals["data_points"]:
            model_name = data_point["model_name"]
            if model_name in figure_x_names:
                last_found = figure_x_names.index(model_name)
            else:
                idx = last_found + 1
                figure_x_names.insert(idx, model_name)
                if model_name in models_metadata and models_metadata[model_name]["display_name"] is not None:
                    figure_x_display_names.insert(idx, models_metadata[model_name]["display_name"])
                else:
                    figure_x_display_names.insert(idx, model_name)
                if model_name in models_metadata:
                    figure_x_parameters.insert(idx, models_metadata[model_name]["parameters"])
                    figure_x_MACs.insert(idx, models_metadata[model_name]["MACs"])
                else:
                    figure_x_parameters.insert(idx, None)
                    figure_x_MACs.insert(idx, None)
                last_found = idx

    colors = plt.rcParams['axes.prop_cycle'].by_key()['color']
    plots_names = []
    plots_colors = []
    plots_markers = []
    plots_y_flash = []
    plots_y_ram = []
    plots_y_exe = []
    plots_y_error = []
    for plot_name, plot_vals in plots.items():
        plots_names.append(plot_name)
        plots_markers.append(plot_vals["marker"])
        if isinstance(plot_vals["color"], float) or isinstance(plot_vals["color"], int):
            plots_colors.append(colors[int(plot_vals["color"])-1])
        else:
            plots_colors.append(plot_vals["color"])

        plot_y_flash = [None] * len(figure_x_names)
        plot_y_ram = [None] * len(figure_x_names)
        plot_y_exe = [None] * len(figure_x_names)
        plot_y_error = [None] * len(figure_x_names)
        for data_point in plot_vals["data_points"]:
            model_name = data_point["model_name"]

            plot_y_flash[figure_x_names.index(model_name)] = data_point["flash"]
            plot_y_ram[figure_x_names.index(model_name)] = data_point["ram"]
            plot_y_exe[figure_x_names.index(model_name)] = data_point["exe"]["average"]
            plot_y_error[figure_x_names.index(model_name)] = data_point["error"]["average"]

        plots_y_flash.append(plot_y_flash)
        plots_y_ram.append(plot_y_ram)
        plots_y_exe.append(plot_y_exe)
        plots_y_error.append(plot_y_error)

    plots_x_parameters_valid = all([param is not None for param in figure_x_parameters])
    plots_x_MACs_valid = all([MAC is not None for MAC in figure_x_MACs])

    # plot models flash memory
    fig, ax = plt.subplots()

    for i in range(len(plots_names)):
        plot_name = plots_names[i]
        if plots_x_parameters_valid:
            plot_x = figure_x_parameters
            plot_y = plots_y_flash[i]
            plot_x, plot_y = zip(*sorted(zip(plot_x, plot_y)))
        else:
            plot_x = figure_x_display_names
            plot_y = plots_y_flash[i]

        line = ax.plot(plot_x, plot_y, label=plot_name, color=plots_colors[i], marker=plots_markers[i], alpha=opacity)
        fill_plot_gaps(ax, line[0])

    if plots_x_parameters_valid:
        ax.set_xscale('log')
        ax.set_xlabel('Parameters')
    else:
        ax.set_xlabel('Models')
    ax.set_yscale('log')
    ax.set_ylabel('Flash Memory [KB]')
    ax.legend(loc='best')

    ax.grid(True, which='major', linestyle='-', linewidth=0.5)
    ax.grid(True, which='minor', linestyle='--', linewidth=0.2)
    ax.xaxis.set_tick_params(rotation=45)

    save_light(fig, os.path.join(save_dir, "flash"))
    save_dark(fig, os.path.join(save_dir, "dark", "flash"))

    # plot models RAM memory
    fig, ax = plt.subplots()

    for i in range(len(plots_names)):
        plot_name = plots_names[i]
        plot_x = figure_x_display_names
        plot_y = plots_y_ram[i]

        line = ax.plot(plot_x, plot_y, label=plot_name, color=plots_colors[i], marker=plots_markers[i], alpha=opacity)
        fill_plot_gaps(ax, line[0])

    ax.set_yscale('log')
    ax.set_xlabel('Models')
    ax.set_ylabel('RAM Memory [KB]')
    ax.legend(loc='best')

    ax.grid(True, which='major', linestyle='-', linewidth=0.5)
    ax.grid(True, which='minor', linestyle='--', linewidth=0.2)
    ax.xaxis.set_tick_params(rotation=45)

    save_light(fig, os.path.join(save_dir, "ram"))
    save_dark(fig, os.path.join(save_dir, "dark", "ram"))

    # plot models execution time
    fig, ax = plt.subplots()

    for i in range(len(plots_names)):
        plot_name = plots_names[i]
        if plots_x_MACs_valid:
            plot_x = figure_x_MACs
            plot_y = plots_y_exe[i]
            plot_x, plot_y = zip(*sorted(zip(plot_x, plot_y)))
        else:
            plot_x = figure_x_display_names
            plot_y = plots_y_exe[i]

        line = ax.plot(plot_x, plot_y, label=plot_name, color=plots_colors[i], marker=plots_markers[i], alpha=opacity)
        fill_plot_gaps(ax, line[0])

    if plots_x_MACs_valid:
        ax.set_xscale('log')
        ax.set_xlabel('MACs')
    else:
        ax.set_xlabel('Models')
    ax.set_yscale('log')
    ax.set_ylabel('Execution Time [ms]')
    ax.legend(loc='best')

    ax.grid(True, which='major', linestyle='-', linewidth=0.5)
    ax.grid(True, which='minor', linestyle='--', linewidth=0.2)
    ax.xaxis.set_tick_params(rotation=45)

    save_light(fig, os.path.join(save_dir, "exe"))
    save_dark(fig, os.path.join(save_dir, "dark", "exe"))

    # plot models error
    fig, ax = plt.subplots()

    for i in range(len(plots_names)):
        plot_name = plots_names[i]
        plot_x = figure_x_display_names
        plot_y = plots_y_error[i]

        line = ax.plot(plot_x, plot_y, label=plot_name, color=plots_colors[i], marker=plots_markers[i], alpha=opacity)
        fill_plot_gaps(ax, line[0])

    ax.set_xlabel('Models')
    ax.set_ylabel('Error [MAE]')
    ax.legend(loc='best')

    ax.grid(True, which='major', linestyle='-', linewidth=0.5)
    ax.grid(True, which='minor', linestyle='--', linewidth=0.2)
    ax.xaxis.set_tick_params(rotation=45)

    save_light(fig, os.path.join(save_dir, "error"))
    save_dark(fig, os.path.join(save_dir, "dark", "error"))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Plot the results of the benchmarking test.")
    parser.add_argument("data_file_path", type=str, help="The path to the excel file containing the benchmarking results.")
    parser.add_argument("save_dir", type=str, help="The directory where the plots will be saved.")
    args = parser.parse_args()

    main(args.data_file_path, args.save_dir)
