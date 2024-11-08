import argparse
import importlib
import os
import re
import shutil
import traceback

import openpyxl
import yaml
from omegaconf import OmegaConf
from openpyxl.styles import Alignment

from edgemark.models.automate.investigator import find_exe_time, find_prediction_mae
from edgemark.models.utils.utils import get_abs_path
from edgemark.models.utils.result_plotter import main as result_plotter


config_file_path = "edgemark/models/automate/configs/automate_config.yaml"
config_file_path = get_abs_path(config_file_path)


class _arena_size_recommender:
    """
    A class to recommend the arena size.
    """
    def __init__(self, initial_arena_size, resolution):
        """
        Initialize the arena size recommender.

        Args:
            initial_arena_size (int): The initial arena size.
            resolution (int): The resolution of search for the arena size.
        """
        self.arena_size = max(round(initial_arena_size / resolution), 1)
        self.resolution = resolution

        self.error_high = None  # the size in which the program gives build error (i.e., the size that is too large)
        self.works = None       # the size in which the program works
        self.error_low = None   # the size in which the program gives runtime error (i.e., the size that is too small)

        self.growth_rate = 1.25
        self.growth_step = 4
        self.shrink_rate = 0.8
        self.shrink_step = 4

    def recommend(self):
        """
        Recommend the next arena size.

        Returns:
            tuple: A tuple containing the recommended size and the status of the search.
                - size (int): The recommended size.
                - status (int): The status of the search. 0 if the arena size is not found yet, 1 if the arena size is found, and -1 if not able to find the arena size.
        """

        if (self.error_high is None) and (self.works is None) and (self.error_low is None):
            return self.arena_size * self.resolution, 0

        if (self.error_high is not None) and (self.works is None) and (self.error_low is None):
            if self.error_high <= 1:
                return None, -1
            candidate_1 = max(1, round(self.error_high * self.shrink_rate))
            next_size = min(candidate_1, self.error_high - self.shrink_step)
            next_size = max(1, next_size)
            return next_size * self.resolution, 0

        if (self.error_high is None) and (self.works is not None) and (self.error_low is None):
            if self.works == 1:
                return self.works * self.resolution, 1
            candidate_1 = max(1, round(self.works * self.shrink_rate))
            next_size = min(candidate_1, self.works - self.shrink_step)
            next_size = max(1, next_size)
            return next_size * self.resolution, 0

        if (self.error_high is None) and (self.works is None) and (self.error_low is not None):
            candidate_1 = max(1, round(self.error_low * self.growth_rate))
            next_size = max(candidate_1, self.error_low + self.growth_step)
            return next_size * self.resolution, 0

        if (self.error_high is not None) and (self.works is not None) and (self.error_low is None):
            if self.works == 1:
                return self.works * self.resolution, 1
            candidate_1 = max(1, round(self.works * self.shrink_rate))
            next_size = min(candidate_1, self.works - self.shrink_step)
            next_size = max(1, next_size)
            return next_size * self.resolution, 0

        if (self.error_high is not None) and (self.works is None) and (self.error_low is not None):
            if self.error_high - self.error_low <= 1:
                return None, -1
            next_size = round((self.error_high + self.error_low) / 2)
            return next_size * self.resolution, 0

        if (self.error_high is None) and (self.works is not None) and (self.error_low is not None):
            if self.works - self.error_low <= 1:
                return self.works * self.resolution, 1
            next_size = round((self.works + self.error_low) / 2)
            return next_size * self.resolution, 0

        if (self.error_high is not None) and (self.works is not None) and (self.error_low is not None):
            if self.works - self.error_low <= 1:
                return self.works * self.resolution, 1
            next_size = round((self.works + self.error_low) / 2)
            return next_size * self.resolution, 0

    def update(self, size, status):
        """
        Update the arena size status.

        Args:
            size (int): The size to update.
            status (int): The status of the size. -1 if the size is too large, 0 if the size works, and +1 if the size is too small.
        """
        assert size % self.resolution == 0, "The size must be a multiple of the resolution. size: {}, resolution: {}".format(size, self.resolution)
        size = size // self.resolution

        if status == -1:
            if self.error_high is None:
                self.error_high = size
            else:
                self.error_high = min(self.error_high, size)

        elif status == 0:
            if self.works is None:
                self.works = size
            else:
                self.works = min(self.works, size)

        elif status == +1:
            if self.error_low is None:
                self.error_low = size
            else:
                self.error_low = max(self.error_low, size)

        else:
            raise ValueError("Invalid status")


def _arena_placer(model_header_path, arena_size):
    """
    Place the arena size in the model header file.

    Args:
        model_header_path (str): The path to the model header file.
        arena_size (int): The arena size to place in the model header file.
    """
    with open(model_header_path, "r") as f:
        lines = f.readlines()

    for i, line in enumerate(lines):
        if "#define ARENA_SIZE" in line:
            lines[i] = "#define ARENA_SIZE {}\n".format(arena_size)
            break

    with open(model_header_path, "w") as f:
        f.writelines(lines)


def _delete_files_in_dir(folder_path):
    """
    Delete all files in the directory, but not the directories themselves.

    Args:
        folder_path (str): The path to the directory
    """
    # This function is a replacement for `shutil.rmtree`.
    # `shutil.rmtree` used to work, but now gives permission error on deleting empty folders in EI library,
    # even when all permissions are granted.
    for dirpath, dirnames, filenames in os.walk(folder_path):
        for filename in filenames:
            file_path = os.path.join(dirpath, filename)
            os.remove(file_path)


def _placer(source_dir, destination_dir, items):
    """
    Place the generated files in the destination directory.

    Args:
        source_dir (str): The directory containing the generated files.
        destination_dir (str): The directory to place the generated files.
        items (list): The list of items to be placed in the destination directory.
    """
    for item in items:
        os.makedirs(destination_dir, exist_ok=True)
        item_path = os.path.join(source_dir, item)
        if os.path.isdir(item_path):
            # if os.path.exists(os.path.join(destination_dir, item)):
            #     shutil.rmtree(os.path.join(destination_dir, item))
            # shutil.copytree(item_path, os.path.join(destination_dir, item))
            if os.path.exists(os.path.join(destination_dir, item)):
                _delete_files_in_dir(os.path.join(destination_dir, item))
            shutil.copytree(item_path, os.path.join(destination_dir, item), dirs_exist_ok=True)
        else:
            if os.path.exists(os.path.join(destination_dir, item)):
                os.remove(os.path.join(destination_dir, item))
            shutil.copy(item_path, os.path.join(destination_dir, item))

        # update the timestamp for "make" to work properly
        for root, dirs, files in os.walk(destination_dir):
            for d in dirs:
                os.utime(os.path.join(root, d), None)
            for f in files:
                os.utime(os.path.join(root, f), None)


def _find_model_name(target_dir):
    """
    Find the model name from the target directory.

    Args:
        target_dir (str): The target directory.

    Returns:
        str: The model name. None if the model name is not found.
    """
    try:
        # keep the part until the date-time string (including date-time string)
        model_base_dir = re.search(r"(.+?)(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})", target_dir).group(0)

        model_info = OmegaConf.load(os.path.join(model_base_dir, "model_info.yaml"))
        setting_file = model_info["setting_file"]
        model_name = os.path.splitext(os.path.basename(setting_file))[0]

        return model_name

    except Exception:
        return None


def _find_model_type(target_dir):
    """
    Find the model type from the target directory.

    Args:
        target_dir (str): The target directory.

    Returns:
        str: The model type. None if the model type is not found.
    """
    pattern = r"(basic|q_dynamic|q_full_int_only|q_full_int|q_16x8_int_only|q_16x8|q_float16|p_\d+|c_\d+)"
    matches = re.findall(pattern, target_dir)

    if matches:
        return " + ".join(matches)
    else:
        return None


def _save_to_excel(benchmarking_result, sheet_name, save_path):
    """
    Save the benchmarking result to an Excel file.

    Args:
        benchmarking_result (list): The list of benchmarking results.
        sheet_name (str): The name of the sheet.
        save_path (str): The path to save the Excel file. Should have the '.xlsx' extension.
    """
    assert save_path.endswith(".xlsx"), "The save path should have the '.xlsx' extension."

    wb = openpyxl.Workbook()

    sheet = wb.active           # select the default sheet
    sheet.title = sheet_name

    # create the title row
    sheet["A1"] = "Model Name"
    sheet.merge_cells("A1:A2")

    sheet["B1"] = "Model Type"
    sheet.merge_cells("B1:B2")

    sheet["C1"] = "Flash (kB)"
    sheet.merge_cells("C1:C2")

    sheet["D1"] = "RAM (kB)"
    sheet.merge_cells("D1:D2")

    sheet["E1"] = "Execution Time (ms)"
    sheet.merge_cells("E1:G1")

    sheet["E2"] = "n tests"
    sheet["F2"] = "average"
    sheet["G2"] = "std"

    sheet["H1"] = "Error (MAE)"
    sheet.merge_cells("H1:J1")

    sheet["H2"] = "n tests"
    sheet["I2"] = "average"
    sheet["J2"] = "std"

    sheet["K1"] = "Model Directory"
    sheet.merge_cells("K1:K2")

    sheet.merge_cells("A3:Z3")

    # fill the data
    data_row = 4

    model_names = set([element["model_name"] for element in benchmarking_result])   # get the unique model names
    # sort the model names and put the None model at the end
    model_names = sorted(model_names, key=lambda x: x if x is not None else "zzz")
    for model_name in model_names:
        model_elements = [element for element in benchmarking_result if element["model_name"] == model_name]

        sheet.cell(row=data_row, column=1, value=model_name)
        if model_name is not None:
            sheet.merge_cells(start_row=data_row, start_column=1, end_row=data_row+len(model_elements)-1, end_column=1)

        for element in model_elements:
            sheet.cell(row=data_row, column=2, value=element["model_type"])
            # 'data' is ignored
            if element["text_size"] is not None:
                sheet.cell(row=data_row, column=3, value=(element["text_size"])/1024)
            if element["bss_size"] is not None:
                sheet.cell(row=data_row, column=4, value=(element["bss_size"])/1024)
            sheet.cell(row=data_row, column=5, value=element["n_timing_tests"])
            sheet.cell(row=data_row, column=6, value=element["avg_ms"])
            sheet.cell(row=data_row, column=7, value=element["std_ms"])
            sheet.cell(row=data_row, column=8, value=element["n_accuracy_tests"])
            sheet.cell(row=data_row, column=9, value=element["avg_mae"])
            sheet.cell(row=data_row, column=10, value=element["std_mae"])
            sheet.cell(row=data_row, column=11, value=element["model_directory"])
            data_row += 1

        sheet.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=26)
        data_row += 1

    # style the sheet
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["K"].width = 20

    for row_num in range(1, data_row):
        row = sheet[row_num]
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    for row_num in [1, 2]:
        row = sheet[row_num]
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A3"   # freeze the title row

    # save the Excel file
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    wb.save(save_path)


def main(cfg_path=config_file_path, **kwargs):
    """
    Compile, upload, and test the models on the hardware device.

    Args:
        cfg_path (str): The path to the configuration file.
            The configuration file that this path points to should contain the following keys:
                - software_platform (str): The software platform to be tested. It should be one of the following: ['TFLM', 'EI', 'Ekkono', 'eAI_Translator']
                - hardware_platform (str): The hardware platform to be tested.
                - linkers_dir (str): The directory containing the linker files.
                - save_dir (str): The directory to save the benchmarking results.
                - benchmark_overall_timeout (float): The overall timeout for reading the benchmark output in seconds.
                - benchmark_silence_timeout (float): The silence timeout for reading the benchmark output in seconds.
        **kwargs (dict): Keyword arguments to be passed to the configuration file.

    Returns:
        list: A list of dictionaries containing the following keys for each target model:
            - dir (str): The directory of the target model.
            - result (str): Result of the model generation. It can be either "success" or "failed".
            - error (str): Error message in case of failure.
            - traceback (str): Traceback in case of failure. Either this or 'error_file' will be present.
            - error_file (str): The path to the error file in case of failure. Either this or 'traceback' will be present.
    """
    cfg = OmegaConf.load(cfg_path)
    cfg.update(OmegaConf.create(kwargs))

    assert cfg.software_platform is not None, "The software platform must be provided."
    assert cfg.software_platform in ["TFLM", "EI", "Ekkono", "eAI_Translator"], "The software platform must be either TFLM, EI, Ekkono, or eAI_Translator."
    assert cfg.hardware_platform is not None, "The hardware platform must be provided."

    spec = importlib.util.spec_from_file_location("imported_module", cfg.hardware_platform_path)
    imported_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(imported_module)
    hardware = imported_module.Hardware(cfg.software_platform)

    if cfg.software_platform == "TFLM":
        targets = OmegaConf.load(os.path.join(cfg.linkers_dir, "tflm_converted_models_list.yaml"))
    elif cfg.software_platform == "EI":
        targets = OmegaConf.load(os.path.join(cfg.linkers_dir, "ei_converted_models_list.yaml"))
    elif cfg.software_platform == "Ekkono":
        targets = OmegaConf.load(os.path.join(cfg.linkers_dir, "ekkono_converted_models_list.yaml"))
    elif cfg.software_platform == "eAI_Translator":
        targets = OmegaConf.load(os.path.join(cfg.linkers_dir, "translator_converted_models_list.yaml"))

        # check if the user has put the eAI_Translator files in the target directories
        eAI_Translator_dirs_exist = True
        for target_dir in targets:
            if not os.path.exists(os.path.join(target_dir, "Translator")):
                eAI_Translator_dirs_exist = False
                break

        if not eAI_Translator_dirs_exist:
            print("Please use the Renesas eAI_Translator and convert TFLite models to eAI_Translator files.")
            print("You can skip a model by creating an empty 'Translator' folder in the target directory.")

        while not eAI_Translator_dirs_exist:
            eAI_Translator_dirs_exist = True
            for target_dir in targets:
                if not os.path.exists(os.path.join(target_dir, "Translator")):
                    eAI_Translator_dirs_exist = False
                    print("")
                    print("TFLite model should be found in: {}".format(target_dir.replace("eAI_Translator", "tflite")))
                    print("eAI_Translator model should be placed in: {}".format(target_dir))

            if not eAI_Translator_dirs_exist:
                print("")
                input("Press Enter when you have put the eAI_Translator files in the target directories ...")

        for target_dir in targets:
            if os.path.exists(os.path.join(target_dir, "Translator")) and not os.listdir(os.path.join(target_dir, "Translator")):
                targets.remove(target_dir)

    if cfg.software_platform == "TFLM":
        replacing_items = ["model.h", "model.cpp", "data.h", "data.cpp"]
    elif cfg.software_platform == "EI":
        replacing_items = ["model", "data.h", "data.cpp"]
    elif cfg.software_platform == "Ekkono":
        replacing_items = ["model.h", "model.c", "data.h", "data.c"]
    elif cfg.software_platform == "eAI_Translator":
        replacing_items = ["Translator", "data.h", "data.c"]

    output = [{"dir": target_dir} for target_dir in targets]

    benchmarking_result = []
    for i, target_dir in enumerate(targets):
        try:
            target_dir = target_dir.replace("\\", "/")
            save_dir = os.path.join(target_dir, "benchmark_result", cfg.software_platform + " + " + cfg.hardware_platform)

            title = "Testing the model in {} ({}/{})".format(target_dir, i+1, len(targets))
            print("\n")
            print("="*110)
            print("-"*((110-len(title)-2)//2), end=" ")
            print(title, end=" ")
            print("-"*((110-len(title)-2)//2))
            print("="*110)

            print("Placing the model's files/folders ...", end=" ", flush=True)
            destination_dir = hardware.get_model_dir()
            _placer(target_dir, destination_dir, replacing_items)

            if cfg.software_platform == "EI":
                if cfg.hardware_platform == "NUCLEO-L4R5ZI":
                    # remove all folders inside "{destination_dir}/model/edge-impulse-sdk/porting" except "stm32-cubeai"
                    for subdir in next(os.walk(os.path.join(destination_dir, "model/edge-impulse-sdk/porting")))[1]:
                        if subdir not in "stm32-cubeai":
                            # shutil.rmtree(os.path.join(destination_dir, "model/edge-impulse-sdk/porting", subdir))
                            _delete_files_in_dir(os.path.join(destination_dir, "model/edge-impulse-sdk/porting", subdir))
                elif cfg.hardware_platform == "RenesasRX65N":
                    # remove all folders inside "{destination_dir}/model/edge-impulse-sdk/porting"
                    for subdir in next(os.walk(os.path.join(destination_dir, "model/edge-impulse-sdk/porting")))[1]:
                        # shutil.rmtree(os.path.join(destination_dir, "model/edge-impulse-sdk/porting", subdir))
                        _delete_files_in_dir(os.path.join(destination_dir, "model/edge-impulse-sdk/porting", subdir))
                    _placer(os.path.dirname(cfg.EI_general_porting_dir), os.path.join(destination_dir, "model/edge-impulse-sdk/porting"), ["general"])

            print("Done")

            # find the best arena_size for the TFLM model. Although it's not general, we'll do it for TFLM :)
            if cfg.arena_finder and cfg.software_platform == "TFLM":
                print("Finding the best arena size ...")

                model_header_path = os.path.join(destination_dir, "model.h")

                original_arena_size = None
                with open(model_header_path, "r") as f:
                    pattern = r"#define ARENA_SIZE (\d+)"
                    matches = re.findall(pattern, f.read())
                    if matches:
                        original_arena_size = int(matches[0])

                if original_arena_size is not None:
                    arena_size = original_arena_size - 10240 + 2048     # reducing 10kB that was added for safety and adding 2kB have a bit of room for breath
                else:
                    arena_size = 4096

                if arena_size < 16384:
                    search_resolution = 512
                elif arena_size < 65536:
                    search_resolution = 1024
                else:
                    search_resolution = 2048

                recommender = _arena_size_recommender(arena_size, search_resolution)

                founded_arena = None
                while True:
                    arena_size, recommender_status = recommender.recommend()

                    if recommender_status == 1:
                        print("The best arena size is found to be: {}".format(arena_size))
                        founded_arena = arena_size
                        _arena_placer(model_header_path, arena_size)
                        break

                    if recommender_status == -1:
                        print("The best arena size cannot be found")
                        _arena_placer(model_header_path, original_arena_size)
                        break

                    print("Building the project with arena size of {} ...".format(arena_size), end=" ", flush=True)
                    _arena_placer(model_header_path, arena_size)
                    try:
                        text_size, data_size, bss_size = hardware.build_project()
                        print("Done")
                    except hardware.RAMExceededError:
                        print("Failed")
                        recommender.update(arena_size, -1)
                        continue
                    except Exception:
                        print("Failed")
                        print("Unknown build error!")
                        _arena_placer(model_header_path, original_arena_size)
                        break

                    print("Uploading the program ...", end=" ", flush=True)
                    try:
                        hardware.upload_program()
                        print("Done")
                    except Exception:
                        print("Failed")
                        _arena_placer(model_header_path, original_arena_size)
                        break

                    print("Reading the output ...", end=" ", flush=True)

                    try:
                        hardware.read_output(overall_timeout=cfg.benchmark_overall_timeout, silence_timeout=cfg.benchmark_silence_timeout, keyword="Benchmark end")
                        recommender.update(arena_size, 0)
                        print("Done")
                    except hardware.BoardNotFoundError:
                        print("Failed")
                        print("Board not found!")
                        _arena_placer(model_header_path, original_arena_size)
                        break
                    except TimeoutError as e:
                        if "missing" in str(e) or "Too many buffers" in str(e):
                            print("Failed")
                            recommender.update(arena_size, 1)
                            continue
                        else:
                            print("Failed")
                            print("Receiving output timeout!")
                            _arena_placer(model_header_path, original_arena_size)
                            break
                    except Exception:
                        print("Failed")
                        print("Unknown error!")
                        _arena_placer(model_header_path, original_arena_size)
                        break
                print("")

            print("Building the project ...", end=" ", flush=True)
            try:
                text_size, data_size, bss_size = hardware.build_project()
                print("Done")

            except hardware.RAMExceededError as e:
                print("Failed")
                shutil.rmtree(save_dir, ignore_errors=True)
                os.makedirs(save_dir, exist_ok=True)
                with open(os.path.join(save_dir, "build_error.txt"), "w") as f:
                    f.write(str(e))
                output[i]["result"] = "failed"
                output[i]["error"] = "RAM size exceeded"
                output[i]["error_file"] = os.path.join(save_dir, "build_error.txt")
                continue

            except hardware.FlashExceededError as e:
                print("Failed")
                shutil.rmtree(save_dir, ignore_errors=True)
                os.makedirs(save_dir, exist_ok=True)
                with open(os.path.join(save_dir, "build_error.txt"), "w") as f:
                    f.write(str(e))
                output[i]["result"] = "failed"
                output[i]["error"] = "Flash size exceeded"
                output[i]["error_file"] = os.path.join(save_dir, "build_error.txt")
                continue

            except Exception as e:
                print("Failed")
                shutil.rmtree(save_dir, ignore_errors=True)
                os.makedirs(save_dir, exist_ok=True)
                with open(os.path.join(save_dir, "build_error.txt"), "w") as f:
                    f.write(str(e))
                output[i]["result"] = "failed"
                output[i]["error"] = "Build failed"
                output[i]["error_file"] = os.path.join(save_dir, "build_error.txt")
                continue

            result = {"text_size": text_size, "data_size": data_size, "bss_size": bss_size}

            if cfg.software_platform == "TFLM":
                result.update({"arena_finder": cfg.arena_finder})
                if cfg.arena_finder:
                    result.update({"arena_resolution": search_resolution, "founded_arena": founded_arena})

            print("Uploading the program ...", end=" ", flush=True)
            try:
                hardware.upload_program()
                print("Done")
            except Exception as e:
                print("Failed")
                shutil.rmtree(save_dir, ignore_errors=True)
                os.makedirs(save_dir, exist_ok=True)
                with open(os.path.join(save_dir, "upload_error.txt"), "w") as f:
                    f.write(str(e))
                output[i]["result"] = "failed"
                output[i]["error"] = "Upload failed"
                output[i]["error_file"] = os.path.join(save_dir, "upload_error.txt")
                continue

            print("Reading the output ...", end=" ", flush=True)

            try:
                benchmark_output = hardware.read_output(overall_timeout=cfg.benchmark_overall_timeout, silence_timeout=cfg.benchmark_silence_timeout, keyword="Benchmark end")
                print("Done")

            except hardware.BoardNotFoundError as e:
                print("Failed")
                shutil.rmtree(save_dir, ignore_errors=True)
                os.makedirs(save_dir, exist_ok=True)
                with open(os.path.join(save_dir, "connection_error.txt"), "w") as f:
                    f.write(str(e))
                output[i]["result"] = "failed"
                output[i]["error"] = "Board not found"
                output[i]["error_file"] = os.path.join(save_dir, "connection_error.txt")
                continue

            except TimeoutError as e:
                print("Failed")
                shutil.rmtree(save_dir, ignore_errors=True)
                os.makedirs(save_dir, exist_ok=True)
                with open(os.path.join(save_dir, "timeout_error.txt"), "w") as f:
                    f.write(str(e))
                output[i]["result"] = "failed"
                if "missing" in str(e) or "Too many buffers" in str(e):
                    output[i]["error"] = "Arena size is too small"
                else:
                    output[i]["error"] = "Receiving output timeout"
                output[i]["error_file"] = os.path.join(save_dir, "timeout_error.txt")
                continue

            except Exception as e:
                print("Failed")
                shutil.rmtree(save_dir, ignore_errors=True)
                os.makedirs(save_dir, exist_ok=True)
                with open(os.path.join(save_dir, "benchmark_output_error.txt"), "w") as f:
                    f.write(str(e))
                output[i]["result"] = "failed"
                output[i]["error"] = "Benchmark output error"
                output[i]["error_file"] = os.path.join(save_dir, "benchmark_output_error.txt")
                continue

            result.update({"serial_output": benchmark_output})

            n_timing_tests, avg_ms, std_ms, avg_ticks, std_ticks = find_exe_time(benchmark_output)
            result.update({
                "n_timing_tests": n_timing_tests,
                "avg_ms": float(avg_ms) if avg_ms is not None else None,
                "std_ms": float(std_ms) if std_ms is not None else None,
                "avg_ticks": float(avg_ticks) if avg_ticks is not None else None,
                "std_ticks": float(std_ticks) if std_ticks is not None else None
            })

            n_accuracy_tests, avg_mae, std_mae = find_prediction_mae(benchmark_output)
            result.update({
                "n_accuracy_tests": n_accuracy_tests,
                "avg_mae": float(avg_mae) if avg_mae is not None else None,
                "std_mae": float(std_mae) if std_mae is not None else None
            })

            benchmarking_element = result.copy()
            benchmarking_element.update({
                "model_name": _find_model_name(target_dir),
                "model_type": _find_model_type(target_dir),
                "model_directory": target_dir
            })
            benchmarking_result.append(benchmarking_element)

            # 'data' section is taken by our samples data and platforms don't affect it
            # (TFLM and Ekkono don't, but EI will slightly affect it which we ignore).
            # So, we can safely remove it from Flash and RAM sizes.
            print("Benchmarking result:")
            print("Flash size: {} bytes".format(text_size))
            print("RAM size: {} bytes".format(bss_size))

            if n_timing_tests > 0:
                print("n_timing_tests: {}".format(n_timing_tests))
                print("avg_ms: {} ms".format(avg_ms))
                print("std_ms: {} ms".format(std_ms))
                print("avg_ticks: {}".format(avg_ticks))
                print("std_ticks: {}".format(std_ticks))
            else:
                print("No timing information was found")

            if n_accuracy_tests > 0:
                print("n_accuracy_tests: {}".format(n_accuracy_tests))
                print("avg_mae: {}".format(avg_mae))
                print("std_mae: {}".format(std_mae))
            else:
                print("No accuracy information was found")

            shutil.rmtree(save_dir, ignore_errors=True)
            os.makedirs(save_dir, exist_ok=True)
            with open(os.path.join(save_dir, "result.yaml"), "w") as f:
                yaml.dump(result, f, indent=4, sort_keys=False)

            output[i]["result"] = "success"

        except Exception as e:
            output[i]["result"] = "failed"
            output[i]["error"] = type(e).__name__
            output[i]["traceback"] = traceback.format_exc()
            print("Error:")
            print(traceback.format_exc())

    _save_to_excel(benchmarking_result, cfg.software_platform, cfg.save_path)

    test_name = os.path.basename(cfg.save_path)
    test_name = os.path.splitext(test_name)[0]
    figures_save_dir = os.path.join(os.path.dirname(cfg.save_path), "figures", test_name)
    result_plotter(cfg.save_path, figures_save_dir)

    return output


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Compile, upload, and test the models on the STM32 board.")
    parser.add_argument("software_platform", type=str, help="The software platform to be tested.")
    parser.add_argument("hardware_platform", type=str, help="The hardware platform to be tested.")
    parser.add_argument("--cfg_path", type=str, default=config_file_path, help="The path to the configuration file.")
    args = parser.parse_args()

    main(args.cfg_path, software_platform=args.software_platform, hardware_platform=args.hardware_platform)
